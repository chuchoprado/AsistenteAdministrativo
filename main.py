import os
import re
import json
import time
import base64
import asyncio
import sqlite3
import logging
import tempfile
import subprocess
from datetime import datetime
from contextlib import closing
from typing import Any, Dict, List, Optional

import fitz  # PyMuPDF
import anthropic
from fastapi import FastAPI, Request
from telegram import Update
from telegram.constants import ChatAction
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

import speech_recognition as sr
from gtts import gTTS
from pydub import AudioSegment
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

app = FastAPI()


# ─────────────────────────────────────────────
# UTILIDADES GENERALES
# ─────────────────────────────────────────────

def convert_oga_to_wav(oga_path: str, wav_path: str) -> bool:
    try:
        subprocess.run(
            ["ffmpeg", "-y", "-i", oga_path, wav_path],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return True
    except Exception as e:
        logger.error(f"Error convirtiendo audio: {e}")
        return False


def safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip()
    if not text:
        return None
    text = (
        text.replace("€", "")
            .replace("EUR", "")
            .replace("Gs.", "")
            .replace("$", "")
            .strip()
            .replace(" ", "")
    )
    if text.count(".") >= 1 and text.count(",") == 1:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    try:
        return round(float(text), 2)
    except Exception:
        return None


def parse_date_to_iso(date_text: Any) -> Optional[str]:
    if not date_text:
        return None
    raw = str(date_text).strip()
    if not raw:
        return None
    patterns = [
        "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y",
        "%d.%m.%Y", "%Y/%m/%d", "%d-%m-%y", "%d/%m/%y",
    ]
    for fmt in patterns:
        try:
            return datetime.strptime(raw, fmt).strftime("%d-%m-%Y")
        except Exception:
            pass
    match = re.search(r"(\d{2})[\/\-.](\d{2})[\/\-.](\d{4})", raw)
    if match:
        d, m, y = match.groups()
        try:
            return datetime(int(y), int(m), int(d)).strftime("%d-%m-%Y")
        except Exception:
            return None
    return None


def normalize_estado(value: Any) -> str:
    text = str(value or "").strip().upper()
    mapping = {
        "COMPLETA": "COMPLETA",
        "OK": "COMPLETA",
        "COMPLETE": "COMPLETA",
        "VERIFICAR_DATOS": "VERIFICAR_DATOS",
        "VERIFICAR": "VERIFICAR_DATOS",
        "REVISAR": "VERIFICAR_DATOS",
        "PENDIENTE_REVISION": "PENDIENTE_REVISION",
        "PENDIENTE": "PENDIENTE_REVISION",
        "NO_PROCESABLE": "PENDIENTE_REVISION",
    }
    return mapping.get(text, "PENDIENTE_REVISION")


def build_fallback_row(page_num: int, reason: str) -> Dict[str, Any]:
    return {
        "numero_factura": None,
        "pagina": page_num,
        "fecha_literal": None,
        "fecha_iso": None,
        "total_eur": None,
        "iva_pct": 10,
        "base_eur": None,
        "cuota_eur": None,
        "estado": "PENDIENTE_REVISION",
        "observaciones": reason[:250] if reason else "No procesable",
    }


def normalize_row(row: Dict[str, Any], page_num: int) -> Dict[str, Any]:
    normalized = {
        "numero_factura": row.get("numero_factura"),
        "pagina": row.get("pagina") if row.get("pagina") is not None else page_num,
        "fecha_literal": row.get("fecha_literal"),
        "fecha_iso": parse_date_to_iso(row.get("fecha_iso") or row.get("fecha_literal")),
        "total_eur": safe_float(row.get("total_eur")),
        "iva_pct": 10,
        "base_eur": None,
        "cuota_eur": None,
        "estado": normalize_estado(row.get("estado")),
        "observaciones": str(row.get("observaciones") or "OK").strip(),
    }

    if normalized["total_eur"] is not None:
        base = round(normalized["total_eur"] / 1.10, 2)
        cuota = round(normalized["total_eur"] - base, 2)
        normalized["base_eur"] = base
        normalized["cuota_eur"] = cuota

    if not normalized["fecha_iso"] and normalized["estado"] == "COMPLETA":
        normalized["estado"] = "VERIFICAR_DATOS"
        if normalized["observaciones"] == "OK":
            normalized["observaciones"] = "Falta fecha confiable"

    if normalized["total_eur"] is None and normalized["estado"] == "COMPLETA":
        normalized["estado"] = "VERIFICAR_DATOS"
        if normalized["observaciones"] == "OK":
            normalized["observaciones"] = "Falta total confiable"

    if (
        normalized["fecha_iso"] is None
        and normalized["total_eur"] is None
        and normalized["estado"] != "PENDIENTE_REVISION"
    ):
        normalized["estado"] = "PENDIENTE_REVISION"

    return normalized


def sort_and_renumber_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def sort_key(item: Dict[str, Any]):
        date_iso = item.get("fecha_iso")
        if date_iso:
            # Convertir DD-MM-YYYY a YYYY-MM-DD para ordenación correcta
            try:
                parts = date_iso.split("-")
                if len(parts) == 3 and len(parts[2]) == 4:
                    sortable = f"{parts[2]}-{parts[1]}-{parts[0]}"
                else:
                    sortable = date_iso
            except Exception:
                sortable = date_iso
            return (0, sortable, item.get("pagina") or 0)
        return (1, "9999-99-99", item.get("pagina") or 0)

    rows_sorted = sorted(rows, key=sort_key)
    for idx, row in enumerate(rows_sorted, start=1):
        row["numero_factura"] = f"Z-{idx}"
    return rows_sorted


# ─────────────────────────────────────────────
# GENERACIÓN DE EXCEL
# ─────────────────────────────────────────────

def create_invoice_excel(rows: List[Dict[str, Any]]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # ── Estilos ──
    header_fill  = PatternFill(fill_type="solid", fgColor="1F3864")
    alt_fill     = PatternFill(fill_type="solid", fgColor="EBF0FA")
    total_fill   = PatternFill(fill_type="solid", fgColor="D9E1F2")
    white_font   = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    bold_font    = Font(bold=True, name="Arial", size=10)
    base_font    = Font(name="Arial", size=10)
    thin_border  = Border(
        left=Side(style="thin", color="B8CCE4"),
        right=Side(style="thin", color="B8CCE4"),
        top=Side(style="thin", color="B8CCE4"),
        bottom=Side(style="thin", color="B8CCE4"),
    )
    med_border = Border(
        left=Side(style="medium", color="1F3864"),
        right=Side(style="medium", color="1F3864"),
        top=Side(style="medium", color="1F3864"),
        bottom=Side(style="medium", color="1F3864"),
    )

    headers = [
        "N° Factura", "Página", "Fecha Literal", "Fecha ISO",
        "Total €", "IVA %", "Base €", "Cuota IVA €", "Estado", "Observaciones",
    ]

    # ── Cabecera (fila 1) ──
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = med_border
    ws.row_dimensions[1].height = 32

    # ── Filas de datos ──
    for row_idx, row in enumerate(rows, start=2):
        ws.append([
            row.get("numero_factura"),
            row.get("pagina"),
            row.get("fecha_literal"),
            row.get("fecha_iso"),
            row.get("total_eur"),
            row.get("iva_pct"),
            row.get("base_eur"),
            row.get("cuota_eur"),
            row.get("estado"),
            row.get("observaciones"),
        ])

        # Color base: filas alternas
        fila_fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type="solid", fgColor="FFFFFF")

        for col_idx in range(1, 11):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font   = base_font
            cell.fill   = fila_fill

        # Color de estado (columna I = 9)
        estado_cell = ws.cell(row=row_idx, column=9)
        estado = str(row.get("estado") or "")
        if estado == "COMPLETA":
            estado_cell.fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
            estado_cell.font = Font(color="375623", bold=True, name="Arial", size=10)
        elif estado == "VERIFICAR_DATOS":
            estado_cell.fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
            estado_cell.font = Font(color="9C5700", bold=True, name="Arial", size=10)
        elif estado == "PENDIENTE_REVISION":
            estado_cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
            estado_cell.font = Font(color="9C0006", bold=True, name="Arial", size=10)

        estado_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Formato numérico y alineación
        for col in [5, 7, 8]:
            c = ws.cell(row=row_idx, column=col)
            c.number_format = '#,##0.00 €'
            c.alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=10).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Fila de TOTALES ──
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTALES").font = bold_font
    for col in range(1, 11):
        c = ws.cell(row=total_row, column=col)
        c.fill   = total_fill
        c.border = med_border
        c.font   = bold_font
    for col in [5, 7, 8]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col)
        c.value        = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        c.number_format = '#,##0.00 €'
        c.alignment    = Alignment(horizontal="right")

    # ── Filtros automáticos y panel fijo ──
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    # ── Anchos de columna ──
    widths = [14, 8, 18, 14, 12, 8, 12, 14, 22, 55]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── Hoja 2: Resumen ──
    ws2 = wb.create_sheet("Resumen")

    total_paginas = len(rows)
    completas  = sum(1 for r in rows if r.get("estado") == "COMPLETA")
    verificar  = sum(1 for r in rows if r.get("estado") == "VERIFICAR_DATOS")
    pendientes = sum(1 for r in rows if r.get("estado") == "PENDIENTE_REVISION")
    vacias     = sum(1 for r in rows if "vac" in str(r.get("observaciones", "")).lower())
    no_proc    = sum(
        1 for r in rows
        if any(x in str(r.get("observaciones", "")).lower()
               for x in ["ilegible", "no procesable", "no factura", "ocr", "error"])
    )
    pct_completas = round((completas / total_paginas) * 100, 2) if total_paginas else 0.0

    if pendientes >= max(2, total_paginas // 2):
        estado_general = "PROBLEMAS_MULTIPLES"
        estado_color   = "FFC7CE"
    elif verificar > 0 or pendientes > 0:
        estado_general = "REQUIERE_REVISION"
        estado_color   = "FFEB9C"
    else:
        estado_general = "EXITOSO"
        estado_color   = "C6EFCE"

    # Cabecera del Resumen
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Resumen de extracción"
    ws2["A1"].fill      = header_fill
    ws2["A1"].font      = white_font
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    summary_rows = [
        ("Total páginas procesadas", total_paginas),
        ("Completas",                completas),
        ("Verificar datos",          verificar),
        ("Pendientes revisión",      pendientes),
        ("Páginas vacías",           vacias),
        ("No procesables",           no_proc),
        ("% Completas",              f"{pct_completas}%"),
        ("Estado general",           estado_general),
    ]

    alt_fill2 = PatternFill(fill_type="solid", fgColor="EBF0FA")
    thin2 = Border(
        left=Side(style="thin", color="B8CCE4"),
        right=Side(style="thin", color="B8CCE4"),
        top=Side(style="thin", color="B8CCE4"),
        bottom=Side(style="thin", color="B8CCE4"),
    )
    for idx, (label, value) in enumerate(summary_rows, start=2):
        c1 = ws2.cell(row=idx, column=1, value=label)
        c2 = ws2.cell(row=idx, column=2, value=value)
        fill = alt_fill2 if idx % 2 == 0 else PatternFill(fill_type="solid", fgColor="FFFFFF")
        for c in [c1, c2]:
            c.border = thin2
            c.fill   = fill
            c.font   = base_font
        c1.font = Font(bold=True, name="Arial", size=10)

    ws2.cell(row=9, column=2).fill = PatternFill(fill_type="solid", fgColor=estado_color)
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22

    # ── Guardar ──
    temp_dir   = tempfile.mkdtemp(prefix="bot_excel_")
    month_year = datetime.now().strftime("%m%Y")
    file_path  = os.path.join(temp_dir, f"Facturas_{month_year}.xlsx")
    wb.save(file_path)
    return file_path


# ─────────────────────────────────────────────
# BOT PRINCIPAL
# ─────────────────────────────────────────────

class CoachBot:
    def __init__(self):
        required_env_vars = {
            "TELEGRAM_TOKEN":   os.getenv("TELEGRAM_TOKEN"),
            "ANTHROPIC_API_KEY": os.getenv("ANTHROPIC_API_KEY"),
            "WEBHOOK_URL":      os.getenv("WEBHOOK_URL"),
        }

        missing_vars = [var for var, value in required_env_vars.items() if not value]
        if missing_vars:
            raise EnvironmentError(
                f"Faltan variables de entorno requeridas: {', '.join(missing_vars)}"
            )

        self.telegram_token = required_env_vars["TELEGRAM_TOKEN"]
        self.webhook_url    = required_env_vars["WEBHOOK_URL"].strip().strip('"').strip("'")

        # Modelo Claude — puedes sobreescribir con la variable ANTHROPIC_MODEL
        self.model  = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-20250514").strip()
        self.client = anthropic.AsyncAnthropic(api_key=required_env_vars["ANTHROPIC_API_KEY"])

        if self.webhook_url.startswith("WEBHOOK_URL="):
            self.webhook_url = self.webhook_url.split("=", 1)[1].strip()
        if not self.webhook_url.startswith("https://"):
            raise EnvironmentError(f"WEBHOOK_URL inválida: {self.webhook_url}")
        if "/webhook" not in self.webhook_url:
            raise EnvironmentError(
                f"WEBHOOK_URL debe apuntar a /webhook. Valor actual: {self.webhook_url}"
            )

        self.started = False
        self.pending_requests: set           = set()
        self.processing_documents: set       = set()
        self.processed_document_ids: Dict[str, float] = {}
        self.processed_update_ids:   Dict[int, float] = {}
        self.update_lock   = asyncio.Lock()
        self.document_lock = asyncio.Lock()

        self.db_path = "bot_data.db"
        self.user_preferences: Dict[int, Dict[str, Any]] = {}
        self.locks: Dict[int, asyncio.Lock] = {}

        self.telegram_app = Application.builder().token(self.telegram_token).build()

        self._init_db()
        self._load_user_preferences()
        self.setup_handlers()

    # ── Anti-duplicados ──

    def _prune_processed_ids(self):
        now = time.time()
        self.processed_update_ids   = {k: v for k, v in self.processed_update_ids.items()   if now - v < 3600}
        self.processed_document_ids = {k: v for k, v in self.processed_document_ids.items() if now - v < 7200}

    async def is_duplicate_update(self, update_id: int) -> bool:
        async with self.update_lock:
            self._prune_processed_ids()
            if update_id in self.processed_update_ids:
                return True
            self.processed_update_ids[update_id] = time.time()
            return False

    async def start_document_processing(self, document_key: str) -> bool:
        async with self.document_lock:
            self._prune_processed_ids()
            if document_key in self.processing_documents:
                return False
            if document_key in self.processed_document_ids:
                return False
            self.processing_documents.add(document_key)
            return True

    async def finish_document_processing(self, document_key: str):
        async with self.document_lock:
            self.processing_documents.discard(document_key)
            self.processed_document_ids[document_key] = time.time()

    async def abort_document_processing(self, document_key: str):
        async with self.document_lock:
            self.processing_documents.discard(document_key)

    # ── Base de datos ──

    def _init_db(self):
        with closing(sqlite3.connect(self.db_path)) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS conversations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    chat_id INTEGER,
                    role TEXT,
                    content TEXT
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS user_preferences (
                    chat_id INTEGER PRIMARY KEY,
                    voice_responses BOOLEAN DEFAULT 0,
                    voice_speed FLOAT DEFAULT 1.0
                )
            """)
            conn.commit()

    def _load_user_preferences(self):
        with closing(sqlite3.connect(self.db_path)) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT chat_id, voice_responses, voice_speed FROM user_preferences")
            for chat_id, voice_responses, voice_speed in cursor.fetchall():
                self.user_preferences[chat_id] = {
                    "voice_responses": bool(voice_responses),
                    "voice_speed": voice_speed,
                }

    def save_conversation(self, chat_id: int, role: str, content: str):
        with closing(sqlite3.connect(self.db_path)) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO conversations (chat_id, role, content) VALUES (?, ?, ?)",
                (chat_id, role, content),
            )
            conn.commit()

    def save_user_preference(
        self,
        chat_id: int,
        voice_responses: Optional[bool] = None,
        voice_speed: Optional[float]    = None,
    ):
        pref = self.user_preferences.get(chat_id, {"voice_responses": False, "voice_speed": 1.0})
        if voice_responses is not None:
            pref["voice_responses"] = voice_responses
        if voice_speed is not None:
            pref["voice_speed"] = voice_speed
        self.user_preferences[chat_id] = pref

        with closing(sqlite3.connect(self.db_path)) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT OR REPLACE INTO user_preferences (chat_id, voice_responses, voice_speed) VALUES (?, ?, ?)",
                (chat_id, int(pref["voice_responses"]), pref["voice_speed"]),
            )
            conn.commit()

    # ── Comandos de voz ──

    async def enable_voice_responses(self, chat_id: int) -> str:
        self.save_user_preference(chat_id, voice_responses=True)
        return "✅ Respuestas por voz activadas."

    async def disable_voice_responses(self, chat_id: int) -> str:
        self.save_user_preference(chat_id, voice_responses=False)
        return "✅ Respuestas por voz desactivadas."

    async def set_voice_speed(self, chat_id: int, text: str) -> str:
        try:
            parts = text.lower().split("velocidad")
            if len(parts) < 2:
                return "⚠️ Indica una velocidad. Ejemplo: velocidad 1.2"
            speed = float(parts[1].strip())
            if speed < 0.5 or speed > 2.0:
                return "⚠️ La velocidad debe estar entre 0.5 y 2.0"
            self.save_user_preference(chat_id, voice_speed=speed)
            return f"✅ Velocidad de voz configurada en {speed}x"
        except ValueError:
            return "⚠️ No entendí la velocidad. Usa algo como 0.8, 1.0 o 1.5"

    async def process_voice_command(self, chat_id: int, text: str) -> Optional[str]:
        text_lower = text.lower()
        if "activar voz"    in text_lower or "activa voz"    in text_lower:
            return await self.enable_voice_responses(chat_id)
        if "desactivar voz" in text_lower or "desactiva voz" in text_lower:
            return await self.disable_voice_responses(chat_id)
        if "velocidad" in text_lower:
            return await self.set_voice_speed(chat_id, text_lower)
        return None

    # ── Llamadas a Claude ──

    async def call_text_model(self, user_message: str) -> str:
        """Responde mensajes de texto genéricos usando Claude."""
        response = await self.client.messages.create(
            model=self.model,
            max_tokens=1024,
            system=(
                "Responde en español de forma útil y directa. "
                "Si el usuario no está pidiendo extraer un PDF, responde normalmente."
            ),
            messages=[
                {"role": "user", "content": user_message}
            ],
        )
        return response.content[0].text.strip()

    async def extract_invoice_from_page(
        self, page_b64: str, page_num: int
    ) -> Dict[str, Any]:
        """
        Extrae los datos de una página de factura usando Claude Vision.
        Recibe el contenido en base64 puro (sin prefijo data:...).
        """
        system_prompt = (
            "Eres un extractor contable experto en tickets y facturas escaneadas o fotografiadas. "
            "Analizas UNA sola imagen de página. "
            "Extrae datos reales visibles. Nunca inventes datos. "
            "Si un valor no puede determinarse con seguridad razonable, usa null. "
            "Busca especialmente la fecha de emisión y el total final pagado. "
            "No confundas subtotal con total final. "
            "Si la página no es procesable, marca estado como PENDIENTE_REVISION.\n\n"
            "INSTRUCCIÓN CRÍTICA: Responde EXCLUSIVAMENTE con un objeto JSON válido. "
            "Sin texto antes ni después. Sin markdown. Sin backticks. Solo el JSON.\n\n"
            "El JSON debe tener exactamente estos campos:\n"
            '{\n'
            '  "numero_factura": null,\n'
            '  "pagina": <entero>,\n'
            '  "fecha_literal": "<texto original o null>",\n'
            '  "fecha_iso": "<DD-MM-YYYY o null>",\n'
            '  "total_eur": <número decimal o null>,\n'
            '  "iva_pct": <número o null>,\n'
            '  "base_eur": null,\n'
            '  "cuota_eur": null,\n'
            '  "estado": "COMPLETA" | "VERIFICAR_DATOS" | "PENDIENTE_REVISION",\n'
            '  "observaciones": "<máx 20 palabras>"\n'
            '}'
        )

        user_prompt = (
            f"Analiza esta imagen correspondiente a la página {page_num} del PDF. "
            "Extrae la fecha de emisión y el total final pagado. "
            "Si hay duda, usa null. "
            "Devuelve ÚNICAMENTE el JSON con los datos reales visibles. "
            "No devuelvas plantillas vacías ni ejemplos."
        )

        try:
            response = await self.client.messages.create(
                model=self.model,
                max_tokens=1024,
                system=system_prompt,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/png",
                                    "data": page_b64,
                                },
                            },
                            {
                                "type": "text",
                                "text": user_prompt,
                            },
                        ],
                    }
                ],
            )

            raw_text = response.content[0].text.strip()

            # Limpiar posibles markdown fences que Claude pueda añadir
            raw_text = re.sub(r"^```(?:json)?\s*", "", raw_text)
            raw_text = re.sub(r"\s*```$", "", raw_text)
            raw_text = raw_text.strip()

            if not raw_text:
                return build_fallback_row(page_num, "Sin respuesta del modelo")

            parsed = json.loads(raw_text)
            if not isinstance(parsed, dict):
                return build_fallback_row(page_num, "Respuesta JSON inválida")

            parsed["pagina"] = page_num
            normalized = normalize_row(parsed, page_num)

            if (
                normalized["fecha_iso"] is None
                and normalized["total_eur"] is None
                and normalized["observaciones"] in ("", "OK")
            ):
                normalized["estado"]       = "PENDIENTE_REVISION"
                normalized["observaciones"] = "No se pudo extraer fecha ni total"

            return normalized

        except json.JSONDecodeError as e:
            logger.error(f"Error JSON página {page_num}: {e} — respuesta: {raw_text[:300]}")
            return build_fallback_row(page_num, f"Error parseando JSON en página {page_num}")
        except anthropic.APIStatusError as e:
            logger.error(f"Error API Anthropic página {page_num}: {e}", exc_info=True)
            return build_fallback_row(page_num, f"Error API en página {page_num}")
        except Exception as e:
            logger.error(f"Error extrayendo página {page_num}: {e}", exc_info=True)
            return build_fallback_row(page_num, f"Error inesperado en página {page_num}")

    # ── Renderizado de PDF en imágenes ──

    def pdf_to_page_images(self, pdf_path: str, dpi: int = 260) -> List[str]:
        """
        Convierte cada página del PDF en base64 puro (PNG).
        Retorna una lista de strings base64 listos para la API de Claude.
        """
        doc = fitz.open(pdf_path)
        pages_b64: List[str] = []
        zoom   = dpi / 72.0
        matrix = fitz.Matrix(zoom, zoom)

        for page_index in range(len(doc)):
            page     = doc.load_page(page_index)
            pix      = page.get_pixmap(matrix=matrix, alpha=False)
            png_bytes = pix.tobytes("png")
            b64      = base64.b64encode(png_bytes).decode("utf-8")
            pages_b64.append(b64)

        doc.close()
        return pages_b64

    # ── Handler de texto ──

    async def process_text_message(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        user_message: str,
    ) -> str:
        chat_id = update.message.chat.id
        lock    = self.locks.setdefault(chat_id, asyncio.Lock())

        async with lock:
            try:
                if not user_message.strip():
                    return "⚠️ No se recibió un mensaje válido."

                voice_cmd = await self.process_voice_command(chat_id, user_message)
                if voice_cmd:
                    return voice_cmd

                await context.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)

                response = await self.call_text_model(user_message)

                self.save_conversation(chat_id, "user",      user_message)
                self.save_conversation(chat_id, "assistant", response)

                return response or "⚠️ No obtuve una respuesta válida."

            except Exception as e:
                logger.error(f"Error en process_text_message: {e}", exc_info=True)
                return "⚠️ Ocurrió un error al procesar tu mensaje."

    # ── Handler de PDF ──

    async def handle_pdf(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        pdf_path     = None
        document_key = None

        try:
            if not update.message or not update.message.document:
                return

            chat_id  = update.message.chat.id
            document = update.message.document

            if document.mime_type != "application/pdf":
                await update.message.reply_text("⚠️ Solo puedo procesar archivos PDF por ahora.")
                return

            document_key = document.file_unique_id or document.file_id

            if not await self.start_document_processing(document_key):
                logger.info(f"Documento duplicado ignorado: {document_key}")
                return

            lock = self.locks.setdefault(chat_id, asyncio.Lock())
            async with lock:
                if chat_id in self.pending_requests:
                    logger.info(f"Solicitud ya en proceso para chat {chat_id}")
                    return

                self.pending_requests.add(chat_id)

                try:
                    await context.bot.send_chat_action(
                        chat_id=chat_id, action=ChatAction.UPLOAD_DOCUMENT
                    )

                    tg_file   = await document.get_file()
                    safe_name = document.file_name or f"archivo_{chat_id}.pdf"
                    pdf_path  = os.path.join(tempfile.gettempdir(), safe_name)
                    await tg_file.download_to_drive(pdf_path)

                    logger.info(f"PDF descargado: {pdf_path}")
                    await update.message.reply_text(
                        "📄 PDF recibido. Procesando páginas con visión IA (Claude)..."
                    )

                    page_images = await asyncio.to_thread(
                        self.pdf_to_page_images, pdf_path, 260
                    )
                    if not page_images:
                        await update.message.reply_text("⚠️ No pude renderizar páginas del PDF.")
                        return

                    extracted_rows: List[Dict[str, Any]] = []
                    total_pages = len(page_images)

                    for idx, page_b64 in enumerate(page_images, start=1):
                        await context.bot.send_chat_action(
                            chat_id=chat_id, action=ChatAction.TYPING
                        )
                        row = await self.extract_invoice_from_page(page_b64, idx)
                        extracted_rows.append(row)

                        if idx == 1 or idx == total_pages or idx % 5 == 0:
                            await update.message.reply_text(
                                f"🔎 Procesadas {idx}/{total_pages} páginas..."
                            )

                    extracted_rows = sort_and_renumber_rows(extracted_rows)
                    excel_path     = create_invoice_excel(extracted_rows)

                    with open(excel_path, "rb") as f:
                        await update.message.reply_document(
                            document=f,
                            filename=os.path.basename(excel_path),
                            caption="✅ Aquí tienes el Excel generado desde la extracción del PDF.",
                        )

                    completas  = sum(1 for r in extracted_rows if r["estado"] == "COMPLETA")
                    verificar  = sum(1 for r in extracted_rows if r["estado"] == "VERIFICAR_DATOS")
                    pendientes = sum(1 for r in extracted_rows if r["estado"] == "PENDIENTE_REVISION")

                    summary_text = (
                        f"📊 Resumen del proceso:\n"
                        f"• Páginas procesadas: {len(extracted_rows)}\n"
                        f"• ✅ Completas: {completas}\n"
                        f"• ⚠️ Verificar: {verificar}\n"
                        f"• ❌ Pendientes: {pendientes}"
                    )
                    await update.message.reply_text(summary_text)

                    # Limpieza de temporales
                    try:
                        os.remove(excel_path)
                        parent = os.path.dirname(excel_path)
                        if os.path.isdir(parent):
                            os.rmdir(parent)
                    except Exception:
                        pass

                    await self.finish_document_processing(document_key)

                finally:
                    self.pending_requests.discard(chat_id)

        except Exception as e:
            logger.error(f"Error procesando PDF: {e}", exc_info=True)
            if update.message:
                await update.message.reply_text("⚠️ Ocurrió un error procesando el PDF.")
            if document_key:
                await self.abort_document_processing(document_key)
        finally:
            try:
                if pdf_path and os.path.exists(pdf_path):
                    os.remove(pdf_path)
            except Exception as e:
                logger.error(f"Error eliminando PDF temporal: {e}")

    # ── Handler de voz ──

    async def handle_voice_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        oga_file_path = None
        wav_file_path = None

        try:
            chat_id    = update.message.chat.id
            voice_file = await update.message.voice.get_file()

            oga_file_path = f"{chat_id}_voice_note.oga"
            wav_file_path = f"{chat_id}_voice_note.wav"

            await voice_file.download_to_drive(oga_file_path)

            if not convert_oga_to_wav(oga_file_path, wav_file_path):
                await update.message.reply_text("⚠️ No se pudo procesar el archivo de audio.")
                return

            recognizer = sr.Recognizer()
            with sr.AudioFile(wav_file_path) as source:
                audio = recognizer.record(source)

            try:
                user_message = recognizer.recognize_google(audio, language="es-ES")
                logger.info(f'Transcripción de voz: "{user_message}"')
                await update.message.reply_text(f'📝 Tu mensaje: "{user_message}"')

                response = await self.process_text_message(update, context, user_message)
                await self.deliver_text_response(update, context, response)

            except sr.UnknownValueError:
                await update.message.reply_text("⚠️ No pude entender la nota de voz.")
            except sr.RequestError as e:
                logger.error(f"Error en SpeechRecognition: {e}")
                await update.message.reply_text(
                    "⚠️ Error en el servicio de reconocimiento de voz."
                )

        except Exception as e:
            logger.error(f"Error manejando mensaje de voz: {e}", exc_info=True)
            await update.message.reply_text("⚠️ Ocurrió un error procesando la nota de voz.")
        finally:
            for fp in (oga_file_path, wav_file_path):
                try:
                    if fp and os.path.exists(fp):
                        os.remove(fp)
                except Exception as e:
                    logger.error(f"Error eliminando temporal {fp}: {e}")

    # ── Text-to-Speech ──

    async def text_to_speech(self, text: str, speed: float = 1.0) -> Optional[str]:
        try:
            temp_dir = os.path.join(os.getcwd(), "temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, f"voice_{int(time.time())}.mp3")

            tts = gTTS(text=text, lang="es")
            tts.save(temp_path)

            if abs(speed - 1.0) > 0.01:
                song     = AudioSegment.from_mp3(temp_path)
                new_song = song.speedup(playback_speed=speed)
                new_song.export(temp_path, format="mp3")

            return temp_path
        except Exception as e:
            logger.error(f"Error en text_to_speech: {e}", exc_info=True)
            return None

    async def deliver_text_response(
        self,
        update: Update,
        context: ContextTypes.DEFAULT_TYPE,
        response: str,
    ):
        chat_id = update.message.chat.id
        pref    = self.user_preferences.get(
            chat_id, {"voice_responses": False, "voice_speed": 1.0}
        )

        if pref["voice_responses"] and len(response) < 3500:
            voice_path = await self.text_to_speech(response, pref["voice_speed"])
            if voice_path and os.path.exists(voice_path):
                try:
                    await context.bot.send_chat_action(
                        chat_id=chat_id, action=ChatAction.RECORD_AUDIO
                    )
                    with open(voice_path, "rb") as audio:
                        await update.message.reply_voice(audio)
                    os.remove(voice_path)
                    return
                except Exception as e:
                    logger.error(f"Error enviando voz: {e}", exc_info=True)

        await update.message.reply_text(response)

    # ── Comandos ──

    async def start_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text(
            "👋 Bienvenido. Envíame texto, voz o un PDF.\n"
            "Los PDFs se procesan página a página con Claude Vision y te entrego un Excel."
        )

    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        help_text = (
            "🤖 Comandos disponibles:\n\n"
            "/start - Iniciar el bot\n"
            "/help  - Mostrar ayuda\n"
            "/voz   - Configurar respuestas por voz\n\n"
            "Funcionalidades:\n"
            "• Mensajes de texto\n"
            "• Notas de voz\n"
            "• PDFs escaneados o fotografiados\n"
            "• Generación automática de Excel\n\n"
            "Modelo activo: Claude (Anthropic)"
        )
        await update.message.reply_text(help_text)

    async def voice_settings_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        chat_id = update.message.chat.id
        pref    = self.user_preferences.get(
            chat_id, {"voice_responses": False, "voice_speed": 1.0}
        )
        status = "activadas" if pref["voice_responses"] else "desactivadas"
        await update.message.reply_text(
            f"🎙 Configuración de voz\n\n"
            f"Estado actual: respuestas de voz {status}\n"
            f"Velocidad actual: {pref['voice_speed']}x\n\n"
            "Comandos:\n"
            "• Activar voz\n"
            "• Desactivar voz\n"
            "• Velocidad 1.2"
        )

    async def route_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        try:
            await self.handle_message(update, context)
        except Exception as e:
            logger.error(f"Error en route_message: {e}", exc_info=True)
            await update.message.reply_text("❌ Ocurrió un error procesando tu mensaje.")

    async def handle_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        chat_id = update.message.chat.id
        try:
            user_message = (update.message.text or "").strip()
            if not user_message:
                return

            response = await asyncio.wait_for(
                self.process_text_message(update, context, user_message),
                timeout=120.0,
            )

            if not response or not response.strip():
                raise ValueError("La respuesta del modelo está vacía")

            await self.deliver_text_response(update, context, response)

        except asyncio.TimeoutError:
            logger.error(f"Timeout procesando mensaje de {chat_id}")
            await update.message.reply_text(
                "⏳ La operación está tomando demasiado tiempo. Inténtalo más tarde."
            )
        except anthropic.APIStatusError as e:
            logger.error(f"Error API Anthropic: {e}", exc_info=True)
            await update.message.reply_text("❌ Hubo un problema con la API de Claude.")
        except anthropic.APIConnectionError as e:
            logger.error(f"Error de conexión Anthropic: {e}", exc_info=True)
            await update.message.reply_text("❌ No se pudo conectar con Claude. Inténtalo de nuevo.")
        except Exception as e:
            logger.error(f"Error inesperado en handle_message: {e}", exc_info=True)
            await update.message.reply_text("⚠️ Ocurrió un error inesperado.")

    # ── Setup de handlers ──

    def setup_handlers(self):
        self.telegram_app.add_handler(CommandHandler("start", self.start_command))
        self.telegram_app.add_handler(CommandHandler("help",  self.help_command))
        self.telegram_app.add_handler(CommandHandler("voz",   self.voice_settings_command))
        self.telegram_app.add_handler(
            MessageHandler(filters.TEXT & ~filters.COMMAND, self.route_message)
        )
        self.telegram_app.add_handler(MessageHandler(filters.VOICE,        self.handle_voice_message))
        self.telegram_app.add_handler(MessageHandler(filters.Document.PDF, self.handle_pdf))
        logger.info("Handlers configurados correctamente")

    async def async_init(self):
        try:
            await self.telegram_app.initialize()

            if not self.started:
                self.started = True
                await self.telegram_app.start()

            try:
                await self.telegram_app.bot.set_webhook(url=self.webhook_url)
                webhook_info = await self.telegram_app.bot.get_webhook_info()
                logger.info(f"Bot inicializado. Webhook: {self.webhook_url}")
                logger.info(f"Webhook info: {webhook_info}")
            except Exception as e:
                logger.error(f"No se pudo configurar el webhook: {e}", exc_info=True)

        except Exception as e:
            logger.error(f"Error en async_init: {e}", exc_info=True)
            raise


# ─────────────────────────────────────────────
# INICIALIZACIÓN
# ─────────────────────────────────────────────

try:
    bot = CoachBot()
except Exception as e:
    logger.error(f"Error crítico inicializando el bot: {e}", exc_info=True)
    raise


async def process_update_background(update: Update):
    try:
        await bot.telegram_app.process_update(update)
    except Exception as e:
        logger.error(f"Error en procesamiento background del update: {e}", exc_info=True)


# ─────────────────────────────────────────────
# ENDPOINTS FASTAPI
# ─────────────────────────────────────────────

@app.get("/")
async def root():
    return {"status": "ok", "message": "Bot activo — Claude (Anthropic)"}


@app.get("/health")
async def health():
    return {"status": "healthy"}


@app.get("/webhook-info")
async def webhook_info():
    try:
        info = await bot.telegram_app.bot.get_webhook_info()
        return info.to_dict()
    except Exception as e:
        logger.error(f"Error obteniendo webhook_info: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}


@app.on_event("startup")
async def startup_event():
    try:
        await bot.async_init()
        logger.info("Aplicación iniciada correctamente")
    except Exception as e:
        logger.error(f"Error al iniciar la aplicación: {e}", exc_info=True)
        raise


@app.post("/webhook")
async def webhook(request: Request):
    try:
        data   = await request.json()
        update = Update.de_json(data, bot.telegram_app.bot)

        update_id = getattr(update, "update_id", None)
        if update_id is not None:
            if await bot.is_duplicate_update(update_id):
                logger.info(f"Update duplicado ignorado: {update_id}")
                return {"status": "ok", "duplicate": True}

        logger.info(f"Update recibido: {json.dumps(data)[:500]}")
        asyncio.create_task(process_update_background(update))
        return {"status": "ok"}

    except Exception as e:
        logger.error(f"Error procesando webhook: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}
